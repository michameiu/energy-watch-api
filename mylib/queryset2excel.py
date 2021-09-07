import copy
import csv

import openpyxl


# y_name(sheets[0])
#     print(sheet.title)
from django.core.files.storage import default_storage
from openpyxl.utils import get_column_letter

from mylib.my_common import get_digitalocean_spaces_download_url


def exportcsv(headers=[], title="Sheet", filename=None, queryset=[],export_csv=False,request=None):
    """
    Example
    filename="test"
    queryset=[{"school_title":"Warugara","count":4}]
    headers=[{"name":"School Title","value":"school_title"},{"name":"Students Count","value":"count"}]
    path=exportcsv(filename=filename,queryset=queryset,headers=headers,title="Schools",export_csv=True,request=None)
    :param headers:
    :param title:
    :param filename:
    :param querset:
    :return:
    """
    path=""
    #####Validate data, assert headers count match an ojbects attributes

    ###Get the totals
    queryset_length = len(queryset)
    headers_length = len(headers)

    ##New workbook
    wb = openpyxl.Workbook()

    ##Create a sheet
    sheet = wb.active
    sheet.title = title

    ###Set the headers

    for k, col in enumerate(headers):
        cell = sheet.cell(row=1, column=k + 1)
        cell.value = col["name"]

        ####Set the size
        if k + 1 <= headers_length:
            sheet.column_dimensions[get_column_letter(k + 1)].width = 20


    ###Copy the headers to match the number of fields in data
    # try:
    fields_length=len([k for k in queryset[0]])
    # except:
    #     fields_length=len(k for k in queryset[0])
    ##Get number of attributes per  row

    myheaders=copy.deepcopy(headers[:fields_length])

    ####Writing the data
    for i,data in enumerate(queryset):
        ###Loop through all the headers
        for j,col in enumerate(myheaders):
            ##i+2 since (i starts at 0, and the row 1 is for headers)
            dt=data[col["value"]]
            # print(dt)
            dat=",".join(list(dt)) if type(dt) in [list,set] else dt
            # print(dt,dat)
            sheet.cell(row=i+2,column=j+1).value=dat

    ##The output filename

    myfilename="%s.%s"%(filename,"csv" if export_csv else ".xlsx")

    ####Temporatyfilename for openxl
    temp_filename="temp_%s"%(myfilename)

    ##Temporarily save the file
    if export_csv:
        with open(temp_filename, 'wb') as f:  # open('test.csv', 'w', newline="") for python 3
            c = csv.writer(f)
            for r in sheet.rows:
                c.writerow([cell.value for cell in r])
    else:
        wb.save(temp_filename)

    with open(temp_filename) as f:
        default_storage.delete('%s'%(myfilename))
        path=default_storage.save('%s'%(myfilename),f)
    url=path
    if request:
        url = get_digitalocean_spaces_download_url(path) #request.build_absolute_uri(location="/media/" + path)
    return url


if __name__ == '__main__':
    pass

